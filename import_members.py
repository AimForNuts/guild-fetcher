import json
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

def show_popup(title, message):
    """Show a popup message."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showinfo(title, message)
    root.destroy()

def create_empty_guild_df(guild_name):
    """Create an empty DataFrame for a guild that wasn't found."""
    return pd.DataFrame({
        "Name": [],
        "Total Level": [],
        "Combat Level": [],
        "Status": []
    })

def find_guild_file(guild_name):
    """Find all guild files that match the base name."""
    matching_files = []
    for file in os.listdir('.'):
        if file.lower().startswith(guild_name.lower()):
            matching_files.append(file)
    return matching_files

def process_guild_file(file_path):
    """Process guild file and return guild data."""
    try:
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin1']
        data = None
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    data = json.load(f)
                break
            except UnicodeDecodeError:
                continue
            except json.JSONDecodeError:
                continue
        
        if data is None:
            print(f"Could not read file {file_path} with any supported encoding")
            return None
            
        # Strategy 1: Try to get detailed information
        if isinstance(data, dict) and 'guild' in data and isinstance(data['guild'], dict) and 'members' in data['guild']:
            members_data = []
            
            # Get list of participants from the first scheduled raid
            participants = set()
            if 'scheduled_raids' in data['guild'] and data['guild']['scheduled_raids']:
                first_raid = data['guild']['scheduled_raids'][0]
                if 'raid' in first_raid and 'participants' in first_raid['raid']:
                    participants = {p['name'] for p in first_raid['raid']['participants']}
            
            for member in data['guild']['members']:
                # Extract combat level from badges
                combat_level = 'N/A'
                total_level = member.get('total_level', 'N/A')
                
                for badge in member.get('badges', []):
                    if 'Combat Lv.' in badge:
                        combat_level = badge.split('Combat Lv. ')[1]
                        break
                
                # Set status based on whether member is in participants list
                name = member.get('name', '')
                status = 'OK' if name in participants else 'X'
                
                members_data.append({
                    'Name': name,
                    'Total Level': total_level,
                    'Combat Level': combat_level,
                    'Status': status
                })
            return members_data
            
        # Strategy 2: Fallback to simple format
        elif isinstance(data, dict) and 'members' in data:
            members_data = []
            for member in data['members']:
                # Convert status: 'participant' -> 'OK', 'missed' -> 'X'
                status = 'OK' if member.get('status') == 'participant' else 'X'
                members_data.append({
                    'Name': member.get('name', ''),
                    'Total Level': 'N/A',
                    'Combat Level': 'N/A',
                    'Status': status
                })
            return members_data
            
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return None

def main():
    try:
        # Get the directory where the executable is located
        if getattr(sys, 'frozen', False):
            # If the application is run as a bundle
            application_path = os.path.dirname(sys.executable)
        else:
            # If the application is run from a Python interpreter
            application_path = os.path.dirname(os.path.abspath(__file__))
        
        # Change to the application directory
        os.chdir(application_path)
        print(f"Working directory: {application_path}")
        print(f"Current directory contents: {os.listdir('.')}")
        
        # Dictionary to store DataFrames for each guild
        guild_dfs = {}
        
        # Process Ironblood guilds
        base_guilds = ['Ironblood', 'Ironblood II', 'Ironblood III', 'Ironblood IV', 'Ironblood V']
        
        for base_guild in base_guilds:
            print(f"\nProcessing {base_guild}...")
            
            # Find all matching files for this guild
            guild_files = find_guild_file(base_guild)
            print(f"Found files for {base_guild}: {guild_files}")
            
            if not guild_files:
                print(f"No files found for {base_guild}")
                # Create empty DataFrame for guilds without files
                guild_dfs[base_guild] = create_empty_guild_df(base_guild)
                continue
            
            # Process each matching file
            for guild_file in guild_files:
                print(f"Processing file: {guild_file}")
                
                # Process the file
                members_data = process_guild_file(guild_file)
                
                if members_data:
                    print(f"Successfully extracted data from {guild_file}")
                    print(f"Number of members found: {len(members_data)}")
                    # Create DataFrame
                    df = pd.DataFrame(members_data)
                    
                    # Store in dictionary using the base guild name
                    guild_dfs[base_guild] = df
                    print(f"Successfully processed {guild_file}")
                    break  # Use the first valid file for each guild
                else:
                    print(f"Failed to process {guild_file}")
        
        print(f"\nTotal guilds processed: {len(guild_dfs)}")
        for guild, df in guild_dfs.items():
            print(f"{guild}: {len(df)} members")
        
        if not guild_dfs:
            error_msg = "No guilds were successfully processed."
            print(f"❌ {error_msg}")
            show_popup("Error", error_msg)
            return
        
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f'guild_members_{timestamp}.xlsx'
        print(f"\nAttempting to create Excel file: {excel_file}")
        
        try:
            # Export to Excel
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # Write each guild's DataFrame to a separate sheet
                for guild_name, df in guild_dfs.items():
                    print(f"Writing sheet for {guild_name}")
                    # Ensure sheet name is valid (max 31 chars, no special chars)
                    sheet_name = guild_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the worksheet
                    ws = writer.sheets[sheet_name]
                    
                    # Add dropdown to Status column only if there are rows
                    if len(df) > 0:
                        status_validation = DataValidation(type="list", formula1='"OK,X"', allow_blank=True)
                        ws.add_data_validation(status_validation)
                        # Ensure we have at least 2 rows (header + 1 data row)
                        end_row = max(2, len(df) + 1)
                        status_validation.add(f'D2:D{end_row}')
                    
                    # Format headers
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Center align all cells
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
            
            print(f"Excel file created successfully at: {os.path.abspath(excel_file)}")
            success_msg = f"✅ Excel file created with {len(guild_dfs)} guild tabs: {excel_file}"
            print(success_msg)
            show_popup("Success", success_msg)
            
        except Exception as excel_error:
            error_msg = f"❌ Error creating Excel file: {str(excel_error)}"
            print(error_msg)
            show_popup("Error", error_msg)
        
    except Exception as e:
        error_msg = f"❌ An unexpected error occurred: {str(e)}"
        print(error_msg)
        show_popup("Error", error_msg)

if __name__ == "__main__":
    main()
